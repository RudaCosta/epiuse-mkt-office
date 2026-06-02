#!/usr/bin/env python3
"""
gerar_planilha_template_em_branco.py — Template XLSX EM BRANCO com guia por área.

Estrutura:
  Aba 1 "Bruna · Data Intelligence"   — em branco + sugestões + guia
  Aba 2 "Duda · Brand Experience"     — em branco + sugestões + guia
  Aba 3 "Isabela · Field Marketing"   — em branco + sugestões + guia
  Aba 4 "Marlison · Business Dev"     — em branco + sugestões + guia
  Aba 5 "Gantt FY26"                  — consolidado em branco

Saída: vault/00-contexto/metas/template-metas-fy26-em-branco.xlsx
Uso:   python scripts/metas/gerar_planilha_template_em_branco.py
"""
import sys, io
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = ROOT / "vault" / "00-contexto" / "metas" / "template-metas-fy26-em-branco.xlsx"

# ── Cores brand EPI-USE Brasil ─────────────────────────────────────────────
NAVY, RED, BLUE, GREY, CREAM, WHITE = "001844", "CD1543", "004B8D", "CFD1D3", "F8F5F0", "FFFFFF"
GREEN, YELLOW, ORANGE, PURPLE, BROWN = "10B981", "FCD34D", "F97316", "C084FC", "A37D57"

# ── Definição das 4 áreas ──────────────────────────────────────────────────
AREAS = [
    {
        "aba":      "Bruna · Data Intelligence",
        "nome":     "Bruna Yamagami",
        "area":     "Marketing Intelligence & CRM",
        "cor":      "60A5FA",  # azul
        "icone":    "🧠",
        "papel":    "Garante que toda decisão de marketing vem de dado real e auditável. Cuida do CRM, lead scoring, attribution e dashboards de funil.",
        "responsabilidades": [
            "CRM higienizado e atualizado (sem duplicados, com campos críticos preenchidos)",
            "Lead scoring funcional (regras claras, scores atualizados, MQL→SQL bem definidos)",
            "Attribution multi-touch (saber qual canal/conteúdo trouxe cada lead)",
            "Dashboards executivos rodando (funil, origem MQL, conversões por canal)",
            "Reports semanais/mensais de performance",
            "Integração das fontes de dado (Apollo, RD, GA4, LinkedIn Analytics)",
        ],
        "sugestoes_metas": [
            "% de leads novos enriquecidos no CRM em até 24h",
            "% de contatos com campos críticos preenchidos (cargo · empresa · email · tel)",
            "# duplicados removidos/mês",
            "Taxa de conversão MQL → SQL",
            "Tempo médio SLA Marketing → Vendas (handoff)",
            "Acurácia do lead scoring (% de SQL com score alto que vira oportunidade)",
            "# dashboards executivos ativos/atualizados",
            "% de atribuição multi-touch cobrindo o funil",
            "# integrações externas funcionando (Apollo · GA4 · RD · LinkedIn)",
            "Frequência dos reports executivos (semanal · quinzenal · mensal)",
        ],
    },
    {
        "aba":      "Duda · Brand Experience",
        "nome":     "Eduarda Hirose (Duda)",
        "area":     "Brand Experience",
        "cor":      PURPLE,
        "icone":    "🎨",
        "papel":    "Dona da identidade visual, do programa EPI-USE Voices, do Inbound (calendário editorial) e da publicação de Cases. Governa o tom de voz da marca.",
        "responsabilidades": [
            "Identidade visual consistente (DESIGN System, brand guidelines aplicados)",
            "EPI-USE Voices funcionando (Voices ativos publicando, kits aprovados)",
            "Calendário editorial semanal (posts agendados, mix de pilares)",
            "Cases de sucesso publicados (Customer Reference)",
            "Aprovações de criativos/posts no SLA",
            "Branding pré-eventos (capa, brindes, decks)",
        ],
        "sugestoes_metas": [
            "# Voices ativos no programa (meta MVP: 5)",
            "Posts/mês por Voice ativo (meta: 2-4)",
            "# Cases novos publicados no ano (meta FY26: 9 · 2 SAP ERP, 4 SuccessFactors, 1 WFS, 1 ServiceNow, 1 Process)",
            "# Empresas Transformadoras gravadas (meta: 6/ano)",
            "Engajamento médio LinkedIn dos posts Voices",
            "# posts colaboradores/mês (meta: 10/mês)",
            "Tempo médio aprovação criativo (SLA)",
            "# kits Profile Optimizer entregues a Voices",
            "% conformidade com brand guide nos materiais novos",
            "Seguidores Instagram (meta FY26: 3.000)",
            "Seguidores LinkedIn EPI-USE Brasil (meta FY26: 15.000)",
        ],
    },
    {
        "aba":      "Isabela · Field Marketing",
        "nome":     "Isabela Carvalho",
        "area":     "Field Marketing & Eventos",
        "cor":      YELLOW,
        "icone":    "📅",
        "papel":    "Executa o calendário de eventos BR + LATAM (30+/ano), gerencia o MDF SAP, ativações, tática do elefante e cobertura de redes pós-evento.",
        "responsabilidades": [
            "Execução 100% do calendário de eventos planejados (BR + LATAM)",
            "Pré-evento: brindes, artes, stand, vídeos prontos",
            "Cobertura ao vivo (LinkedIn, Instagram) com SLA 24h",
            "Pós-evento: relatório, leads tabulados, reuniões agendadas",
            "MDF SAP — gestão e prestação de contas",
            "Eventos próprios EPI-USE (mínimo 2/ano)",
            "Tática Elefante (giveaway com seguidores LinkedIn ganhos por evento)",
        ],
        "sugestoes_metas": [
            "# eventos executados vs planejados (meta: 100%)",
            "# eventos próprios EPI-USE/ano (meta FY26: 2)",
            "# leads capturados em eventos (mensal)",
            "# reuniões agendadas pós-evento",
            "Custo médio por lead em eventos (R$)",
            "ROI eventos próprios (R$ gerado / R$ investido)",
            "Seguidores ganhos via tática elefante por evento",
            "# reconhecimentos SAP (meta FY26: 4)",
            "# reconhecimentos analistas — ISG/Gartner/Forrester/IDC (meta FY26: 2)",
            "% utilização do MDF SAP no trimestre",
            "Tempo médio cobertura pós-evento (meta SLA: 24h)",
        ],
    },
    {
        "aba":      "Marlison · Business Dev",
        "nome":     "Marlison Estrela",
        "area":     "Development Sales (SDR/BDR)",
        "cor":      "F87171",  # red
        "icone":    "📞",
        "papel":    "Outbound via Apollo, prospecção C-Level (CIOs, CHROs, CFOs), gestão de sequências de email/cold call e bridge Marketing → Vendas.",
        "responsabilidades": [
            "Sequências Apollo rodando (cold email + follow-ups)",
            "Pesquisa de ICP e enriquecimento de contas-alvo",
            "Cold call estruturado (script, ritmo, gravação)",
            "Reuniões agendadas com decisores",
            "CRM atualizado com cada touchpoint (notas, próximos passos)",
            "Handoff qualificado pra time de vendas",
            "Reporting semanal de cadência",
        ],
        "sugestoes_metas": [
            "# contatos novos prospectados/mês (meta FY26: 60 leads/mês)",
            "# sequências ativas no Apollo (hoje: 15 cadastradas, 1 ativa)",
            "# emails enviados/dia",
            "# ligações realizadas/mês (meta FY26: 40/mês)",
            "# emails de outbound/mês (meta FY26: 200/mês)",
            "Taxa de resposta (open rate · reply rate · meeting rate)",
            "# reuniões agendadas/mês (meta FY26: 8/mês)",
            "# oportunidades geradas/mês (meta FY26: 5 ops/mês)",
            "# vendas/ano com origem Marketing/SDR (meta FY26: 10)",
            "Pipeline R$ originado pelo SDR (acumulado YTD)",
            "% CRM atualizado dentro do SLA",
        ],
    },
]

# ── Estilos ────────────────────────────────────────────────────────────────
def estilo_titulo(cell, cor_bg=NAVY):
    cell.font = Font(name="Maven Pro", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=cor_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def estilo_sub(cell, cor_bg=CREAM, cor_txt="0F172A"):
    cell.font = Font(name="Calibri", size=11, italic=True, color=cor_txt)
    cell.fill = PatternFill("solid", fgColor=cor_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

def estilo_section(cell, cor_bg=RED):
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=cor_bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def estilo_header_tabela(cell, cor_bg=NAVY):
    cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=cor_bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=GREY),
        right=Side(style="thin", color=GREY),
        top=Side(style="thin", color=GREY),
        bottom=Side(style="medium", color=RED),
    )

def estilo_celula_vazia(cell):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

def estilo_helper(cell):
    cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
    cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)

# ── Gera 1 aba por pessoa ──────────────────────────────────────────────────
def gerar_aba_pessoa(wb, dados):
    ws = wb.create_sheet(dados["aba"])

    # Larguras de coluna
    widths = [38, 14, 12, 12, 14, 16, 12, 16, 26, 38]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── BLOCO TÍTULO ────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = f"{dados['icone']}  Metas FY26 — {dados['area']}"
    estilo_titulo(ws["A1"], dados["cor"])
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Responsável: {dados['nome']}  ·  Preencher: 1 linha por meta na tabela abaixo"
    estilo_sub(ws["A2"])
    ws.row_dimensions[2].height = 24

    # ── BLOCO PAPEL DA ÁREA ─────────────────────────────────────
    ws.merge_cells("A4:J4")
    ws["A4"] = "🎯  PAPEL DA ÁREA"
    estilo_section(ws["A4"])
    ws.row_dimensions[4].height = 22

    ws.merge_cells("A5:J5")
    ws["A5"] = dados["papel"]
    estilo_helper(ws["A5"])
    ws.row_dimensions[5].height = 38

    # ── BLOCO RESPONSABILIDADES ─────────────────────────────────
    ws.merge_cells("A7:J7")
    ws["A7"] = "✅  RESPONSABILIDADES (escopo da função)"
    estilo_section(ws["A7"])
    ws.row_dimensions[7].height = 22

    for i, item in enumerate(dados["responsabilidades"]):
        r = 8 + i
        ws.merge_cells(f"A{r}:J{r}")
        ws[f"A{r}"] = f"   •  {item}"
        estilo_helper(ws[f"A{r}"])
        ws.row_dimensions[r].height = 22

    # ── BLOCO SUGESTÕES DE METAS ────────────────────────────────
    next_row = 8 + len(dados["responsabilidades"]) + 1
    ws.merge_cells(f"A{next_row}:J{next_row}")
    ws[f"A{next_row}"] = "💡  SUGESTÕES DE METAS (use como inspiração — não copia tudo, escolha o que faz sentido)"
    estilo_section(ws[f"A{next_row}"], YELLOW)
    ws[f"A{next_row}"].font = Font(name="Calibri", size=11, bold=True, color="92400E")
    ws.row_dimensions[next_row].height = 22

    for i, sug in enumerate(dados["sugestoes_metas"]):
        r = next_row + 1 + i
        ws.merge_cells(f"A{r}:J{r}")
        ws[f"A{r}"] = f"   →  {sug}"
        estilo_helper(ws[f"A{r}"])
        ws[f"A{r}"].font = Font(name="Calibri", size=10, italic=True, color="78350F")
        ws[f"A{r}"].fill = PatternFill("solid", fgColor="FEF3C7")
        ws.row_dimensions[r].height = 20

    # ── BLOCO COMO PREENCHER ────────────────────────────────────
    next_row = next_row + len(dados["sugestoes_metas"]) + 2
    ws.merge_cells(f"A{next_row}:J{next_row}")
    ws[f"A{next_row}"] = "📝  COMO PREENCHER A TABELA ABAIXO"
    estilo_section(ws[f"A{next_row}"], BLUE)
    ws.row_dimensions[next_row].height = 22

    instrucoes = [
        "1.  Cada linha = 1 meta (escreva o indicador na 1ª coluna)",
        "2.  Use as sugestões acima OU crie suas próprias metas — você é a/o dona/o desta área",
        "3.  'Meta (valor)' é o número-alvo; 'Unidade' é R$/%/leads/posts/etc; 'Período' = mensal/trimestral/anual",
        "4.  'Realizado YTD' você atualiza periodicamente (mensal de preferência)",
        "5.  '% atingido' é fórmula automática (Realizado / Meta anual)",
        "6.  'Status' tem dropdown: a iniciar / em andamento / atingido / atrasado / cancelado",
        "7.  'Fonte do dado' = de onde sai o número (Apollo / GA4 / LinkedIn / CRM / etc)",
        "8.  Linhas extras: adicione mais embaixo se precisar de mais de 15 metas",
    ]
    for i, txt in enumerate(instrucoes):
        r = next_row + 1 + i
        ws.merge_cells(f"A{r}:J{r}")
        ws[f"A{r}"] = txt
        estilo_helper(ws[f"A{r}"])
        ws.row_dimensions[r].height = 18

    # ── TABELA EM BRANCO ────────────────────────────────────────
    next_row = next_row + len(instrucoes) + 3
    ws.merge_cells(f"A{next_row}:J{next_row}")
    ws[f"A{next_row}"] = "📊  PREENCHA SUAS METAS AQUI"
    estilo_section(ws[f"A{next_row}"], dados["cor"])
    ws[f"A{next_row}"].font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    ws.row_dimensions[next_row].height = 26

    header_row = next_row + 1
    headers = ["Indicador / Meta", "Meta (valor)", "Unidade", "Período",
               "Meta anual", "Realizado YTD", "% atingido", "Status",
               "Fonte do dado", "Observação"]
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=label)
        estilo_header_tabela(c, dados["cor"])
    ws.row_dimensions[header_row].height = 38

    # 15 linhas em branco
    for i in range(15):
        r = header_row + 1 + i
        for col in range(1, 11):
            c = ws.cell(row=r, column=col)
            estilo_celula_vazia(c)
        # Fórmula % atingido (col G = 7) = Realizado (F=6) / Meta anual (E=5)
        ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/E{r},\"\")")
        ws.cell(row=r, column=7).number_format = "0.0%"
        ws.cell(row=r, column=2).number_format = "#,##0.00"
        ws.cell(row=r, column=5).number_format = "#,##0.00"
        ws.cell(row=r, column=6).number_format = "#,##0.00"
        ws.row_dimensions[r].height = 30

    ult = header_row + 15
    # Validação Status
    dv = DataValidation(type="list",
        formula1='"📝 a iniciar,🟡 em andamento,🟢 atingido,🔴 atrasado,🚫 cancelado"',
        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H{header_row+1}:H{ult}")

    # Formatação condicional %
    ws.conditional_formatting.add(f"G{header_row+1}:G{ult}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(f"G{header_row+1}:G{ult}",
        CellIsRule(operator="between", formula=["0.5", "0.999"],
                   fill=PatternFill("solid", fgColor="FEF3C7")))
    ws.conditional_formatting.add(f"G{header_row+1}:G{ult}",
        CellIsRule(operator="lessThan", formula=["0.5"],
                   fill=PatternFill("solid", fgColor="FEE2E2")))

    ws.freeze_panes = f"A{header_row+1}"
    return ws

# ── Aba Gantt consolidada (em branco) ──────────────────────────────────────
def gerar_aba_gantt(wb):
    ws = wb.create_sheet("Gantt FY26 (consolidado)")
    meses = ["Jul/26","Ago/26","Set/26","Out/26","Nov/26","Dez/26",
             "Jan/27","Fev/27","Mar/27","Abr/27","Mai/27","Jun/27"]

    fixos = [("Iniciativa / Meta", 44), ("Área", 22), ("Dona", 18),
             ("Início", 10), ("Fim", 10), ("Status", 14)]

    # Título
    total_cols = len(fixos) + len(meses)
    last_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "📊  Gantt FY26 — Visão temporal Jul/2026 → Jun/2027"
    estilo_titulo(ws["A1"], NAVY)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Cada linha = 1 iniciativa/meta. Pinte manualmente as células dos meses pra refletir duração (Cor de Preenchimento)."
    estilo_sub(ws["A2"])
    ws.row_dimensions[2].height = 22

    # Headers
    for col, (label, w) in enumerate(fixos, start=1):
        c = ws.cell(row=4, column=col, value=label)
        estilo_header_tabela(c)
        ws.column_dimensions[get_column_letter(col)].width = w

    for j, mes in enumerate(meses):
        col = len(fixos) + 1 + j
        c = ws.cell(row=4, column=col, value=mes)
        estilo_header_tabela(c)
        ws.column_dimensions[get_column_letter(col)].width = 8

    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "G5"

    # 30 linhas em branco
    for i in range(30):
        r = 5 + i
        for col in range(1, total_cols + 1):
            estilo_celula_vazia(ws.cell(row=r, column=col))
        ws.row_dimensions[r].height = 24

    # Validação Status na coluna F (6)
    dv = DataValidation(type="list",
        formula1='"📝 a iniciar,🟡 em andamento,🟢 atingido,🔴 atrasado,🚫 cancelado"',
        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F5:F{4+30}")

    # Validação Área (dropdown com as 6 áreas)
    dv_area = DataValidation(type="list",
        formula1='"🧠 Marketing Intelligence,🚀 Growth & Performance,📅 Field Marketing,📞 Development Sales,🎨 Brand Experience,📣 Conteúdo/Redatoria"',
        allow_blank=True)
    ws.add_data_validation(dv_area)
    dv_area.add(f"B5:B{4+30}")

    # Legenda
    legenda_row = 5 + 30 + 2
    ws.cell(row=legenda_row, column=1, value="💡 DICA:").font = Font(bold=True, color=NAVY, size=11)
    dicas = [
        "Selecione as células dos meses (ex: Out/26 → Dez/26) e use 'Cor de Preenchimento' pra pintar a barra do Gantt",
        "Use a cor da área pra manter consistência visual:",
        "    🧠 Intelligence = azul claro    🚀 Growth = verde    📅 Eventos = amarelo",
        "    📞 Pipeline = vermelho    🎨 Brand = roxo    📣 Conteúdo = marrom",
    ]
    for i, txt in enumerate(dicas, start=1):
        ws.cell(row=legenda_row + i, column=1, value=txt).font = Font(size=10, color="475569")

    return ws

# ── Aba ÍNDICE (primeira) ──────────────────────────────────────────────────
def gerar_aba_indice(wb):
    ws = wb.create_sheet("📍 Comece aqui", 0)
    ws.column_dimensions["A"].width = 90

    ws["A1"] = "🎯  TEMPLATE DE METAS FY26 — EPI-USE BRASIL MARKETING"
    estilo_titulo(ws["A1"], NAVY)
    ws.row_dimensions[1].height = 42

    ws["A3"] = "Como usar este arquivo:"
    ws["A3"].font = Font(name="Calibri", size=13, bold=True, color=NAVY)

    instrucoes = [
        "",
        "1.  Cada pessoa abre A SUA aba e preenche as metas da área dela:",
        "       • Bruna  →  aba 'Bruna · Data Intelligence'",
        "       • Duda   →  aba 'Duda · Brand Experience'",
        "       • Isabela →  aba 'Isabela · Field Marketing'",
        "       • Marlison →  aba 'Marlison · Business Dev'",
        "",
        "2.  Cada aba tem:",
        "       • Papel da área  (o que ela faz)",
        "       • Responsabilidades  (escopo)",
        "       • Sugestões de metas em amarelo  (inspiração — não copia tudo)",
        "       • Como preencher  (instruções)",
        "       • Tabela em branco com 15 linhas  (sua entrega)",
        "",
        "3.  A aba 'Gantt FY26' é a visão temporal consolidada:",
        "       • Cole as metas mais importantes (uma por linha)",
        "       • Pinte as células dos meses pra mostrar quando cada iniciativa roda",
        "",
        "4.  Período fiscal: Jul/2026 → Jun/2027  (FY27 começando agora)",
        "",
        "5.  Cores de status (formatação automática):",
        "       🟢 verde   ≥ 100% atingido",
        "       🟡 amarelo entre 50% e 99%",
        "       🔴 vermelho < 50%",
        "",
        "6.  Dúvidas?  Fala com o Rudá.",
    ]
    for i, txt in enumerate(instrucoes, start=4):
        c = ws.cell(row=i, column=1, value=txt)
        if txt.startswith(("1.","2.","3.","4.","5.","6.")):
            c.font = Font(name="Calibri", size=11, bold=True, color=NAVY)
        else:
            c.font = Font(name="Calibri", size=11, color="0F172A")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[i].height = 22

    rodape_row = 4 + len(instrucoes) + 2
    ws.cell(row=rodape_row, column=1,
            value="EPI-USE Brasil · Marketing · Versão 1.0 (Jun/2026) · Geração automática via scripts/metas/").font = \
                Font(name="Calibri", size=9, italic=True, color="94A3B8")

    return ws

# ── Main ───────────────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # Remove a sheet default vazia
    wb.remove(wb.active)

    gerar_aba_indice(wb)
    for area in AREAS:
        gerar_aba_pessoa(wb, area)
    gerar_aba_gantt(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"✅ Salvo: {OUT_PATH}")
    print(f"   {size_kb:.1f} KB · 6 abas (Índice + 4 pessoas + Gantt)")

if __name__ == "__main__":
    main()
