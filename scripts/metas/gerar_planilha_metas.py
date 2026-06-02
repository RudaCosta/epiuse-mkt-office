#!/usr/bin/env python3
"""
gerar_planilha_metas.py — Gera template XLSX de metas FY26 com 2 abas:
  1. "Metas FY26" — linha por meta com colunas pra preenchimento manual
  2. "Gantt FY26" — visão temporal Jul/2026 → Jun/2027 (12 meses)

Lê: public/api/metas-fy26.json (29 metas estruturadas do docx oficial)
Saída: vault/00-contexto/metas/template-metas-fy26.xlsx

Uso: python scripts/metas/gerar_planilha_metas.py
"""
import json
import sys
import io
from pathlib import Path
from datetime import date

# Fix Windows cp1252 console encoding pra suportar emojis
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
METAS_JSON = ROOT / "public" / "api" / "metas-fy26.json"
OUT_PATH = ROOT / "vault" / "00-contexto" / "metas" / "template-metas-fy26.xlsx"

# ── Cores brand EPI-USE Brasil (Brand Guide V1.1) ──────────────────────────
NAVY = "001844"
RED = "CD1543"
BLUE = "004B8D"
GREY_LIGHT = "CFD1D3"
CREAM = "F8F5F0"
WHITE = "FFFFFF"
GREEN = "10B981"
YELLOW = "FCD34D"
ORANGE = "F97316"

# ── Mapa area-do-docx → dona (do agentes.json) ─────────────────────────────
AREA_DONA = {
    "Aquisição & Demanda":          ("📞 Pipeline / SDR",          "Marlison Estrela"),
    "Branding":                      ("🎨 Brand / Voices",          "Eduarda (Duda)"),
    "Conteúdo":                      ("📣 Conteúdo / Redatoria",    "Lisiane de Assis"),
    "Eventos & Branding":            ("📅 Field Marketing",         "Isabela Carvalho"),
    "Programas Especiais":           ("🎨 Brand / Voices",          "Eduarda (Duda)"),
    "Relacionamento":                ("🎨 Brand / Voices",          "Eduarda (Duda)"),
    "Site / Tráfego":                ("🚀 Growth & Performance",    "Guilherme (Gui)"),
    "MQL/SQL":                       ("🧠 Marketing Intelligence",  "Bruna Yamagami"),
    "Outros":                        ("🎯 Rudá / CEO",              "Rudá Costa"),
}

AREA_COLOR = {
    "🧠 Marketing Intelligence":  "60A5FA",
    "🚀 Growth & Performance":     "10B981",
    "📅 Field Marketing":          "FBBF24",
    "📞 Pipeline / SDR":           "F87171",
    "🎨 Brand / Voices":           "C084FC",
    "📣 Conteúdo / Redatoria":     "A37D57",
    "🎯 Rudá / CEO":               "FBBF24",
}

# ── Período FY (12 meses) ──────────────────────────────────────────────────
MESES = [
    ("Jul/26", 7, 2026),  ("Ago/26", 8, 2026),  ("Set/26", 9, 2026),
    ("Out/26", 10, 2026), ("Nov/26", 11, 2026), ("Dez/26", 12, 2026),
    ("Jan/27", 1, 2027),  ("Fev/27", 2, 2027),  ("Mar/27", 3, 2027),
    ("Abr/27", 4, 2027),  ("Mai/27", 5, 2027),  ("Jun/27", 6, 2027),
]

def carregar_metas():
    if not METAS_JSON.exists():
        print(f"AVISO: {METAS_JSON} nao encontrado. Usando lista exemplo.")
        return []
    with open(METAS_JSON, encoding="utf-8") as f:
        d = json.load(f)
    return d.get("metas", [])

# ─────────────────────────────────────────────────────────────────────────────
def estilo_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=GREY_LIGHT),
        right=Side(style="thin", color=GREY_LIGHT),
        top=Side(style="thin", color=GREY_LIGHT),
        bottom=Side(style="medium", color=RED),
    )

def estilo_data(cell, area=None):
    cell.font = Font(name="Calibri", size=10, color="0F172A")
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )
    if area and area in AREA_COLOR:
        cell.fill = PatternFill("solid", fgColor="F8F5F0")

# ─────────────────────────────────────────────────────────────────────────────
def aba_metas(wb, metas):
    ws = wb.active
    ws.title = "Metas FY26"

    headers = [
        ("Área",           18),
        ("Dona",           20),
        ("Indicador",      36),
        ("Meta (valor)",   14),
        ("Unidade",        12),
        ("Período",        12),
        ("Meta anual",     14),
        ("Realizado YTD",  16),
        ("% atingido",     14),
        ("Status",         14),
        ("Fonte do dado",  26),
        ("Observação",     40),
    ]

    for col, (label, w) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=label)
        estilo_header(c)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"

    for i, meta in enumerate(metas, start=2):
        area_raw = meta.get("area", "Outros")
        area_office, dona = AREA_DONA.get(area_raw, (area_raw, "—"))
        valor = meta.get("valor", "")
        valor_ano = meta.get("valor_ano", valor)
        unidade = meta.get("unidade", "")
        periodo = meta.get("periodo", "mensal")

        row = [
            area_office, dona,
            meta.get("label", "(sem nome)"),
            valor, unidade, periodo,
            valor_ano,
            "",  # realizado YTD — preencher manual
            "",  # % atingido — fórmula abaixo
            "📝 a iniciar",
            meta.get("fonte", ""),
            "",
        ]
        for col, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=col, value=val)
            estilo_data(c, area_office)
            if col == 1:
                c.fill = PatternFill("solid", fgColor=AREA_COLOR.get(area_office, "FFFFFF"))
                c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Fórmula % atingido = realizado / meta_anual
        ws.cell(row=i, column=9, value=f"=IFERROR(H{i}/G{i},\"\")")
        ws.cell(row=i, column=9).number_format = "0.0%"
        ws.cell(row=i, column=4).number_format = "#,##0"
        ws.cell(row=i, column=7).number_format = "#,##0"
        ws.cell(row=i, column=8).number_format = "#,##0"

        ws.row_dimensions[i].height = 32

    # Formatação condicional na coluna % atingido
    ult = len(metas) + 1
    if ult > 1:
        ws.conditional_formatting.add(
            f"I2:I{ult}",
            CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                       fill=PatternFill("solid", fgColor="DCFCE7"))
        )
        ws.conditional_formatting.add(
            f"I2:I{ult}",
            CellIsRule(operator="between", formula=["0.5", "0.999"],
                       fill=PatternFill("solid", fgColor="FEF3C7"))
        )
        ws.conditional_formatting.add(
            f"I2:I{ult}",
            CellIsRule(operator="lessThan", formula=["0.5"],
                       fill=PatternFill("solid", fgColor="FEE2E2"))
        )

    # Validação de Status
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list",
        formula1='"📝 a iniciar,🟡 em andamento,🟢 atingido,🔴 atrasado,🚫 cancelado"',
        allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"J2:J{ult}")

    return ws

# ─────────────────────────────────────────────────────────────────────────────
def aba_gantt(wb, metas):
    ws = wb.create_sheet("Gantt FY26")

    # Headers fixos + 12 colunas de mês
    fixos = [
        ("Iniciativa / Meta",   42),
        ("Área",                18),
        ("Dona",                18),
        ("Início (M/AAAA)",     14),
        ("Fim (M/AAAA)",        14),
        ("Status",              14),
    ]
    for col, (label, w) in enumerate(fixos, start=1):
        c = ws.cell(row=1, column=col, value=label)
        estilo_header(c)
        ws.column_dimensions[get_column_letter(col)].width = w

    for j, (mes_label, _, _) in enumerate(MESES):
        col = len(fixos) + 1 + j
        c = ws.cell(row=1, column=col, value=mes_label)
        estilo_header(c)
        ws.column_dimensions[get_column_letter(col)].width = 8

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "G2"  # congela até a coluna Status

    # Adiciona 1 linha por meta (Gantt vazio pra preencher) + algumas iniciativas-exemplo
    for i, meta in enumerate(metas, start=2):
        area_raw = meta.get("area", "Outros")
        area_office, dona = AREA_DONA.get(area_raw, (area_raw, "—"))
        ws.cell(row=i, column=1, value=meta.get("label", "(sem nome)"))
        ws.cell(row=i, column=2, value=area_office)
        ws.cell(row=i, column=3, value=dona)
        ws.cell(row=i, column=4, value="Jul/26")
        ws.cell(row=i, column=5, value="Jun/27")
        ws.cell(row=i, column=6, value="📝 a iniciar")

        for col in range(1, 7):
            estilo_data(ws.cell(row=i, column=col), area_office)
            if col == 2:
                ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=AREA_COLOR.get(area_office, "FFFFFF"))
                ws.cell(row=i, column=2).font = Font(name="Calibri", size=10, bold=True, color=WHITE)
                ws.cell(row=i, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Pinta TODAS as células de mês com a cor da área (representa duração padrão FY)
        cor = AREA_COLOR.get(area_office, "60A5FA")
        for j in range(12):
            col = len(fixos) + 1 + j
            c = ws.cell(row=i, column=col, value="")
            c.fill = PatternFill("solid", fgColor=cor)
            c.border = Border(
                left=Side(style="thin", color="E2E8F0"),
                right=Side(style="thin", color="E2E8F0"),
                top=Side(style="thin", color="E2E8F0"),
                bottom=Side(style="thin", color="E2E8F0"),
            )

        ws.row_dimensions[i].height = 26

    # Legenda de áreas no rodapé
    ult = len(metas) + 4
    ws.cell(row=ult, column=1, value="LEGENDA — áreas").font = Font(bold=True, size=11, color=NAVY)
    for k, (area, cor) in enumerate(AREA_COLOR.items()):
        r = ult + 1 + k
        c = ws.cell(row=r, column=1, value=area)
        c.fill = PatternFill("solid", fgColor=cor)
        c.font = Font(color=WHITE, bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Instruções
    inst_row = ult + len(AREA_COLOR) + 3
    ws.cell(row=inst_row, column=1, value="📌 Como usar:").font = Font(bold=True, color=NAVY, size=11)
    instrucoes = [
        "1. Edite a coluna 'Início' e 'Fim' (formato M/AAAA) pra ajustar o período de cada meta",
        "2. Pinte/limpe manualmente as células dos meses pra refletir a duração real (use ferramenta de cor)",
        "3. Atualize 'Status' conforme o trimestre avança",
        "4. Linhas extras podem ser adicionadas no final pra novas iniciativas dentro do FY",
    ]
    for i, txt in enumerate(instrucoes, start=1):
        ws.cell(row=inst_row + i, column=1, value=txt).font = Font(size=10, color="475569")

    return ws

# ─────────────────────────────────────────────────────────────────────────────
def aba_resumo_areas(wb, metas):
    """Aba 3 (bônus): resumo de quantas metas por área."""
    ws = wb.create_sheet("Resumo por área")
    headers = [("Área", 28), ("Dona", 24), ("Nº de metas", 14), ("Categorias", 50)]
    for col, (label, w) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=label)
        estilo_header(c)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 36

    by_area = {}
    for m in metas:
        area_raw = m.get("area", "Outros")
        area_office, dona = AREA_DONA.get(area_raw, (area_raw, "—"))
        by_area.setdefault(area_office, {"dona": dona, "count": 0, "cats": set()})
        by_area[area_office]["count"] += 1
        by_area[area_office]["cats"].add(m.get("categoria", m.get("label", "")))

    for i, (area, dados) in enumerate(sorted(by_area.items(), key=lambda x: -x[1]["count"]), start=2):
        ws.cell(row=i, column=1, value=area)
        ws.cell(row=i, column=2, value=dados["dona"])
        ws.cell(row=i, column=3, value=dados["count"])
        ws.cell(row=i, column=4, value=" · ".join(sorted(dados["cats"])))
        for col in range(1, 5):
            estilo_data(ws.cell(row=i, column=col))
        ws.cell(row=i, column=1).fill = PatternFill("solid", fgColor=AREA_COLOR.get(area, "FFFFFF"))
        ws.cell(row=i, column=1).font = Font(bold=True, color=WHITE)
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 28

    return ws

# ─────────────────────────────────────────────────────────────────────────────
def main():
    metas = carregar_metas()
    print(f"📊 {len(metas)} metas carregadas de {METAS_JSON.name}")

    wb = Workbook()
    aba_metas(wb, metas)
    aba_gantt(wb, metas)
    aba_resumo_areas(wb, metas)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"✅ Salvo em: {OUT_PATH}")
    print(f"   {size_kb:.1f} KB · 3 abas (Metas FY26 · Gantt FY26 · Resumo por área)")

if __name__ == "__main__":
    main()
