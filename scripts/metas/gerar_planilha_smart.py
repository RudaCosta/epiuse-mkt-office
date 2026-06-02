#!/usr/bin/env python3
"""
gerar_planilha_smart.py — Template SMART em branco + Gantt automático.

Cada pessoa preenche SUAS metas seguindo SMART (Specific · Measurable · Achievable
· Relevant · Time-bound). O Gantt consolidado puxa as datas das abas das pessoas
e pinta sozinho os meses correspondentes via formatação condicional.

Ano fiscal: Mar/2026 → Fev/2027.

Abas:
  📍 Comece aqui     — guia SMART + estrutura
  Bruna              — Data Intelligence (em branco)
  Duda               — Brand Experience (em branco)
  Isabela            — Field Marketing (em branco)
  Marlison           — Business Development (em branco)
  Gantt FY26/27      — auto-consolidado, pinta meses sozinho

Uso: python scripts/metas/gerar_planilha_smart.py
"""
import sys, io
from pathlib import Path
from datetime import date

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = ROOT / "vault" / "00-contexto" / "metas" / "template-metas-fy26-smart.xlsx"

# ── Brand EPI-USE ────────────────────────────────────────────────────────
NAVY, RED, BLUE, GREY, CREAM, WHITE = "001844", "CD1543", "004B8D", "CFD1D3", "F8F5F0", "FFFFFF"
YELLOW_BG, BLUE_BG, PURPLE_BG, RED_BG = "DBEAFE", "FEF3C7", "F3E8FF", "FEE2E2"

# ── Pessoas (nome de aba SIMPLES pra fórmulas funcionarem sem apóstrofos) ──
PESSOAS = [
    {"aba": "Bruna",    "nome": "Bruna Yamagami",  "area": "Marketing Intelligence & CRM (Data Intelligence)", "icone": "🧠", "cor": "60A5FA", "cor_fill": "DBEAFE"},
    {"aba": "Duda",     "nome": "Eduarda Hirose",   "area": "Brand Experience",                                  "icone": "🎨", "cor": "C084FC", "cor_fill": "F3E8FF"},
    {"aba": "Isabela",  "nome": "Isabela Carvalho", "area": "Field Marketing & Eventos",                         "icone": "📅", "cor": "F59E0B", "cor_fill": "FEF3C7"},
    {"aba": "Marlison", "nome": "Marlison Estrela", "area": "Business Development (SDR)",                        "icone": "📞", "cor": "EF4444", "cor_fill": "FEE2E2"},
]

# ── Ano fiscal Mar/26 → Fev/27 ────────────────────────────────────────────
MESES_FY = [
    ("Mar/26", 3, 2026), ("Abr/26", 4, 2026), ("Mai/26", 5, 2026),
    ("Jun/26", 6, 2026), ("Jul/26", 7, 2026), ("Ago/26", 8, 2026),
    ("Set/26", 9, 2026), ("Out/26", 10, 2026), ("Nov/26", 11, 2026),
    ("Dez/26", 12, 2026), ("Jan/27", 1, 2027), ("Fev/27", 2, 2027),
]

# Quantas linhas em branco por pessoa (linha = 1 meta SMART)
LINHAS_POR_PESSOA = 12

# ── Estilos ────────────────────────────────────────────────────────────────
def f_titulo(cell, bg=NAVY):
    cell.font = Font(name="Maven Pro", size=15, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def f_subt(cell, bg=CREAM, color="0F172A"):
    cell.font = Font(name="Calibri", size=10, italic=True, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)

def f_section(cell, bg=RED):
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def f_th(cell, bg=NAVY, color=WHITE):
    cell.font = Font(name="Calibri", size=10, bold=True, color=color)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color=GREY),
        right=Side(style="thin", color=GREY),
        top=Side(style="thin", color=GREY),
        bottom=Side(style="medium", color=RED),
    )

def f_td(cell):
    cell.font = Font(name="Calibri", size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

def f_helper(cell):
    cell.font = Font(name="Calibri", size=10, color="475569")
    cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)


# ═════════════════════════════════════════════════════════════════════════
# ABA 1: COMECE AQUI — guia SMART
# ═════════════════════════════════════════════════════════════════════════
def aba_indice(wb):
    ws = wb.create_sheet("📍 Comece aqui", 0)
    ws.column_dimensions["A"].width = 95

    ws["A1"] = "🎯  METAS FY26 — EPI-USE BRASIL MARKETING"
    f_titulo(ws["A1"])
    ws.row_dimensions[1].height = 42

    ws["A2"] = "Ano fiscal: Março/2026 → Fevereiro/2027 · Use método SMART em cada meta"
    f_subt(ws["A2"])
    ws.row_dimensions[2].height = 24

    # ── Bloco SMART ─────────────────────────────────────────────
    ws["A4"] = "🧭  MÉTODO SMART — Como escrever cada meta"
    f_section(ws["A4"], NAVY)
    ws.row_dimensions[4].height = 26

    smart_blocos = [
        ("S", "Specific (Específico)",
         "O QUE será feito e QUEM está envolvido. Sem ambiguidade.",
         "Ex: 'Lançar campanha LinkedIn Ads pra captar 500 leads C-level de Indústria SAP'", "FCD34D", "92400E"),
        ("M", "Measurable (Mensurável)",
         "NÚMERO + FONTE de onde sai esse número. Sem fonte = não é meta, é desejo.",
         "Ex: '500 leads (Apollo) · CTR ≥ 1,2% (LinkedIn Campaign Manager) · CPL ≤ R$ 80 (Zoho)'", "60A5FA", "1E3A8A"),
        ("A", "Achievable (Atingível)",
         "Os RECURSOS pra entregar existem? Time · budget · ferramentas · aprovações.",
         "Ex: 'Budget R$ 40k aprovado · Agência X contratada · Acesso ao Sales Navigator'", "10B981", "065F46"),
        ("R", "Relevant (Relevante)",
         "POR QUÊ faz sentido pro negócio agora? Conecta com qual prioridade estratégica?",
         "Ex: 'Suporta meta de Roberto de 10 vendas/ano com origem MKT'", "C084FC", "581C87"),
        ("T", "Time-bound (Temporal)",
         "DATA DE INÍCIO + DATA DE FIM. Sem prazo = nunca fica pronto.",
         "Ex: 'Início: 01/abr/2026 · Fim: 30/jun/2026 (sprint Q1 do FY)'", "EF4444", "7F1D1D"),
    ]
    for i, (letra, titulo, descr, exemplo, bg, fg) in enumerate(smart_blocos):
        r = 5 + i * 2
        # Letra grande + título
        ws[f"A{r}"] = f"   {letra}  ·  {titulo}  —  {descr}"
        ws[f"A{r}"].font = Font(name="Calibri", size=11, bold=True, color=fg)
        ws[f"A{r}"].fill = PatternFill("solid", fgColor=bg)
        ws[f"A{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws.row_dimensions[r].height = 26
        # Exemplo
        ws[f"A{r+1}"] = f"          {exemplo}"
        ws[f"A{r+1}"].font = Font(name="Calibri", size=10, italic=True, color="475569")
        ws[f"A{r+1}"].alignment = Alignment(horizontal="left", vertical="center", indent=2)
        ws.row_dimensions[r+1].height = 20

    # ── Bloco FONTES ─────────────────────────────────────────────
    next_row = 5 + len(smart_blocos) * 2 + 1
    ws[f"A{next_row}"] = "📊  FONTES DE DADO (use nomes consistentes — facilita o filtro depois)"
    f_section(ws[f"A{next_row}"], BLUE)
    ws.row_dimensions[next_row].height = 26

    fontes = [
        "Google Analytics 4 · Search Console · LinkedIn Campaign Manager · LinkedIn Analytics (página)",
        "Apollo · Zoho CRM · HubSpot · RD Station · Pipedrive",
        "Microsoft Forms · Google Forms · SharePoint · OneDrive · planilha interna",
        "Formalização por email (com data, remetente, link da thread)",
        "Reports manuais (com link do arquivo · responsável pela atualização · frequência)",
    ]
    for i, fonte in enumerate(fontes):
        r = next_row + 1 + i
        ws[f"A{r}"] = f"   •  {fonte}"
        f_helper(ws[f"A{r}"])
        ws.row_dimensions[r].height = 20

    # ── Como usar ─────────────────────────────────────────────
    next_row = next_row + len(fontes) + 2
    ws[f"A{next_row}"] = "▶️  COMO USAR ESTE ARQUIVO"
    f_section(ws[f"A{next_row}"], RED)
    ws.row_dimensions[next_row].height = 26

    passos = [
        "1.  Cada pessoa abre A SUA aba (Bruna · Duda · Isabela · Marlison)",
        "2.  Preenche 1 meta por linha seguindo SMART. Coluna por coluna.",
        "3.  Datas de Início e Fim DEVEM ser preenchidas (sem isso, Gantt não pinta).",
        "4.  Quando você atualiza o 'Realizado', o '% atingido' calcula sozinho.",
        "5.  A aba 'Gantt FY26/27' consolida TUDO automaticamente — não precisa preencher nada lá.",
        "6.  Ano fiscal: Mar/2026 → Fev/2027 (12 meses). Datas fora desse range aparecem no Gantt sem barra.",
        "7.  Dúvida sobre SMART ou fonte? Fala com o Rudá.",
    ]
    for i, txt in enumerate(passos):
        r = next_row + 1 + i
        ws[f"A{r}"] = txt
        f_helper(ws[f"A{r}"])
        ws.row_dimensions[r].height = 22

    return ws


# ═════════════════════════════════════════════════════════════════════════
# ABAS DE PESSOA — em branco, com headers SMART
# ═════════════════════════════════════════════════════════════════════════
SMART_HEADERS = [
    # (label, largura)
    ("S  ·  Specific\n(o que será feito + quem)",                    44),
    ("M  ·  Measurable\n(número-alvo)",                              16),
    ("Unidade\n(R$ · % · leads · etc)",                              14),
    ("Fonte do dado\n(GA4 · Apollo · Zoho · email · etc)",           26),
    ("A  ·  Achievable\n(recursos necessários)",                     28),
    ("R  ·  Relevant\n(por quê pro negócio)",                        30),
    ("T  ·  Início\n(data)",                                         13),
    ("T  ·  Fim\n(data)",                                            13),
    ("Realizado YTD",                                                15),
    ("% atingido",                                                   13),
    ("Status",                                                       15),
    ("Observação",                                                   34),
]

def aba_pessoa(wb, pessoa):
    ws = wb.create_sheet(pessoa["aba"])

    # Larguras
    for col, (_, w) in enumerate(SMART_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Título ─────────────────────────────────────────────
    last_col = get_column_letter(len(SMART_HEADERS))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{pessoa['icone']}  Metas FY26 — {pessoa['area']}"
    f_titulo(ws["A1"], pessoa["cor"])
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = f"Responsável: {pessoa['nome']}  ·  Preencha 1 linha por meta seguindo método SMART (ver aba 'Comece aqui')"
    f_subt(ws["A2"])
    ws.row_dimensions[2].height = 24

    # ── Lembrete SMART ─────────────────────────────────────
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"] = "🧭  SMART: Specific (o quê + quem) · Measurable (número + FONTE obrigatória) · Achievable (recursos) · Relevant (por quê) · Time-bound (Início + Fim)"
    ws["A4"].font = Font(name="Calibri", size=10, italic=True, bold=True, color="92400E")
    ws["A4"].fill = PatternFill("solid", fgColor="FEF3C7")
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[4].height = 36

    # ── Header da tabela ─────────────────────────────────────────────
    HEADER_ROW = 6
    for col, (label, _) in enumerate(SMART_HEADERS, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=label)
        f_th(c, pessoa["cor"])
    ws.row_dimensions[HEADER_ROW].height = 56

    # ── Linhas em branco ──────────────────────────────────────────────
    first_row = HEADER_ROW + 1
    last_row = HEADER_ROW + LINHAS_POR_PESSOA

    for i in range(LINHAS_POR_PESSOA):
        r = first_row + i
        for col in range(1, len(SMART_HEADERS) + 1):
            c = ws.cell(row=r, column=col)
            f_td(c)
        # Fórmulas e formatos
        # col 7 = Início (data) · col 8 = Fim (data) · col 9 = Realizado · col 10 = %
        ws.cell(row=r, column=2).number_format = "#,##0.00"     # Measurable
        ws.cell(row=r, column=7).number_format = "dd/mm/yyyy"    # Início
        ws.cell(row=r, column=8).number_format = "dd/mm/yyyy"    # Fim
        ws.cell(row=r, column=9).number_format = "#,##0.00"     # Realizado
        ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/B{r},\"\")")  # %
        ws.cell(row=r, column=10).number_format = "0.0%"
        ws.row_dimensions[r].height = 50

    # Dropdown Status (col 11)
    dv_status = DataValidation(type="list",
        formula1='"📝 a iniciar,🟡 em andamento,🟢 atingido,🔴 atrasado,🚫 cancelado"',
        allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"K{first_row}:K{last_row}")

    # Dropdown Fonte (col 4) com sugestões comuns
    dv_fonte = DataValidation(type="list",
        formula1='"GA4,Search Console,LinkedIn Campaign Manager,LinkedIn Analytics,Apollo,Zoho CRM,HubSpot,RD Station,Microsoft Forms,Google Forms,SharePoint,Email (formalizado),Planilha interna,Outro"',
        allow_blank=True)
    ws.add_data_validation(dv_fonte)
    dv_fonte.add(f"D{first_row}:D{last_row}")

    # Formatação condicional na coluna % atingido (col J = 10)
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=PatternFill("solid", fgColor="DCFCE7")))
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
        CellIsRule(operator="between", formula=["0.5", "0.999"],
                   fill=PatternFill("solid", fgColor="FEF3C7")))
    ws.conditional_formatting.add(f"J{first_row}:J{last_row}",
        CellIsRule(operator="lessThan", formula=["0.5"],
                   fill=PatternFill("solid", fgColor="FEE2E2")))

    ws.freeze_panes = f"B{first_row}"
    return ws


# ═════════════════════════════════════════════════════════════════════════
# ABA GANTT — auto-consolidada
# ═════════════════════════════════════════════════════════════════════════
def aba_gantt(wb):
    ws = wb.create_sheet("Gantt FY26-27")

    # Layout: A=Meta, B=Pessoa, C=Início, D=Fim, E..P=12 meses
    fixos = [("Meta (puxa da aba da pessoa)", 50), ("Pessoa", 14), ("Início", 13), ("Fim", 13)]
    total_cols = len(fixos) + len(MESES_FY)
    last_col = get_column_letter(total_cols)

    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "📊  Gantt FY26/27 — Visão consolidada Mar/26 → Fev/27"
    f_titulo(ws["A1"], NAVY)
    ws.row_dimensions[1].height = 42

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = "Esta aba é AUTOMÁTICA. As barras aparecem sozinhas quando a pessoa preenche Início e Fim na aba dela. Não edite as fórmulas."
    f_subt(ws["A2"])
    ws.row_dimensions[2].height = 24

    # Headers
    HEADER_ROW = 4
    for col, (label, w) in enumerate(fixos, start=1):
        c = ws.cell(row=HEADER_ROW, column=col, value=label)
        f_th(c)
        ws.column_dimensions[get_column_letter(col)].width = w

    for j, (mes_label, _, _) in enumerate(MESES_FY):
        col = len(fixos) + 1 + j
        c = ws.cell(row=HEADER_ROW, column=col, value=mes_label)
        f_th(c)
        ws.column_dimensions[get_column_letter(col)].width = 9

    ws.row_dimensions[HEADER_ROW].height = 36
    ws.freeze_panes = f"E{HEADER_ROW + 1}"

    # ── Linhas: 4 blocos (1 por pessoa), com fórmulas referenciando ─────────
    PESSOA_ROW_OFFSET = 7  # primeira linha de dados na aba da pessoa (HEADER_ROW=6 lá, +1)

    current_row = HEADER_ROW + 1
    for pessoa in PESSOAS:
        bloco_inicio = current_row
        for i in range(LINHAS_POR_PESSOA):
            r_pessoa = PESSOA_ROW_OFFSET + i   # linha correspondente na aba da pessoa
            r = current_row

            # Col A: meta (Specific da aba da pessoa, col A=1)
            ws.cell(row=r, column=1, value=f"='{pessoa['aba']}'!A{r_pessoa}")
            f_td(ws.cell(row=r, column=1))

            # Col B: nome da pessoa
            c_pess = ws.cell(row=r, column=2, value=pessoa["nome"].split()[0])
            f_td(c_pess)
            c_pess.fill = PatternFill("solid", fgColor=pessoa["cor"])
            c_pess.font = Font(color=WHITE, bold=True, size=10)
            c_pess.alignment = Alignment(horizontal="center", vertical="center")

            # Col C: Início (data) puxado da aba da pessoa (col G=7)
            cI = ws.cell(row=r, column=3, value=f"='{pessoa['aba']}'!G{r_pessoa}")
            cI.number_format = "dd/mm/yyyy"
            f_td(cI)

            # Col D: Fim (data) puxado (col H=8)
            cF = ws.cell(row=r, column=4, value=f"='{pessoa['aba']}'!H{r_pessoa}")
            cF.number_format = "dd/mm/yyyy"
            f_td(cF)

            # Cols E..P: cada mês
            for j, (_, m, y) in enumerate(MESES_FY):
                col = len(fixos) + 1 + j
                # Fórmula: se Início ≤ último dia do mês E Fim ≥ primeiro dia do mês → "█" senão ""
                # DATE(y,m,1) = início do mês · EOMONTH(DATE(y,m,1),0) = fim do mês
                formula = (
                    f'=IF(AND(ISNUMBER($C{r}),ISNUMBER($D{r}),'
                    f'$C{r}<=EOMONTH(DATE({y},{m},1),0),'
                    f'$D{r}>=DATE({y},{m},1)),"█","")'
                )
                c = ws.cell(row=r, column=col, value=formula)
                f_td(c)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(name="Calibri", size=11, color=pessoa["cor"])

            ws.row_dimensions[r].height = 26
            current_row += 1

        # Formatação condicional pra ESSE bloco da pessoa: célula com "█" recebe cor de fundo
        bloco_fim = current_row - 1
        col_start = get_column_letter(len(fixos) + 1)
        col_end = get_column_letter(total_cols)
        ws.conditional_formatting.add(
            f"{col_start}{bloco_inicio}:{col_end}{bloco_fim}",
            FormulaRule(
                formula=[f'{col_start}{bloco_inicio}<>""'],
                fill=PatternFill("solid", fgColor=pessoa["cor"]),
                font=Font(color=pessoa["cor"], size=11, bold=True),
            )
        )

    # Legenda
    leg_row = current_row + 2
    ws.cell(row=leg_row, column=1, value="LEGENDA — Pessoas").font = Font(bold=True, color=NAVY, size=11)
    for k, p in enumerate(PESSOAS):
        r = leg_row + 1 + k
        c1 = ws.cell(row=r, column=1, value=f"{p['icone']}  {p['nome']}  ·  {p['area']}")
        c1.font = Font(size=10, color="0F172A")
        c2 = ws.cell(row=r, column=2, value="exemplo")
        c2.fill = PatternFill("solid", fgColor=p["cor"])
        c2.font = Font(color=WHITE, bold=True, size=10)
        c2.alignment = Alignment(horizontal="center", vertical="center")

    return ws


# ═════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default

    aba_indice(wb)
    for pessoa in PESSOAS:
        aba_pessoa(wb, pessoa)
    aba_gantt(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    size_kb = OUT_PATH.stat().st_size / 1024
    print(f"✅ Salvo: {OUT_PATH}")
    print(f"   {size_kb:.1f} KB · 6 abas (Índice + Bruna + Duda + Isabela + Marlison + Gantt)")
    print(f"   FY: Mar/2026 → Fev/2027 · {LINHAS_POR_PESSOA} linhas/pessoa · {LINHAS_POR_PESSOA*4} totais no Gantt")

if __name__ == "__main__":
    main()
