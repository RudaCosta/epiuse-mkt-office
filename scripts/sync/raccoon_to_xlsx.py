# Escreve um item do Raccoon na CELULA DO DIA do grid mensal da planilha da Duda (OneDrive).
# Preserva a formatacao (openpyxl). DRY-RUN por padrao: mostra a celula alvo, NAO grava.
# Use --apply para gravar de verdade. Espelha o parser de scripts/sync/sync_calendario_duda.js.
#
# Uso:
#   python raccoon_to_xlsx.py --date 2026-06-16 --titulo "Reforma tributaria: o que o CFO precisa saber" [--apply]
#   python raccoon_to_xlsx.py --from-prod            (puxa itens fonte=raccoon do prod e escreve todos)
import sys, argparse, json, urllib.request, datetime
try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl nao instalado. Rode: python -m pip install openpyxl")

XLSX_PATH = r"C:\Users\Ruds\OneDrive - EPI USE BRASIL SERVICOS EM SISTEMAS LTDA\MARKETING\Inbound\Conteudo\CALENDARIO EDITORIAL EPI-USE (1).xlsx"
# fallback com acento correto (igual ao sync_calendario_duda.js)
XLSX_PATH_ACENTO = r"C:\Users\Ruds\OneDrive - EPI USE BRASIL SERVI" + "Ç" + r"OS EM SISTEMAS LTDA\MARKETING\Inbound\Conte" + "ú" + r"do\CALEND" + "Á" + r"RIO EDITORIAL EPI-USE (1).xlsx"
PROD = "https://office.epiuse.com.br"

MESES = {1:"JANEIRO",2:"FEVEREIRO",3:["MARÇO","MARCO"],4:"ABRIL",5:"MAIO",6:"JUNHO",
         7:"JULHO",8:"AGOSTO",9:"SETEMBRO",10:"OUTUBRO",11:"NOVEMBRO",12:"DEZEMBRO"}

def resolve_path(custom=None):
    import os
    for p in ([custom] if custom else []) + [XLSX_PATH_ACENTO, XLSX_PATH]:
        if p and os.path.exists(p):
            return p
    return None

def month_names(m):
    v = MESES[m]
    return v if isinstance(v, list) else [v]

def find_day_cell(ws, month, day):
    """Acha (row, col) da celula de CONTEUDO do dia no bloco do mes. Espelha parseGrid."""
    names = month_names(month)
    max_row = ws.max_row
    max_col = min(ws.max_column, 12)
    # 1) acha a linha do cabecalho do mes
    month_row = None
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            val = ws.cell(r, c).value
            if val and any(n in str(val).upper() for n in names):
                month_row = r; break
        if month_row: break
    if not month_row:
        return None, None, "mes nao encontrado no grid"
    # 2) a partir de month_row+2: pares (numeros, conteudo)
    r = month_row + 2
    while r <= max_row:
        num_cells = [ws.cell(r, c).value for c in range(1, 8)]
        has_num = any(str(v).strip().isdigit() and 1 <= int(v) <= 31 for v in num_cells if v is not None)
        if not has_num:
            if r > month_row + 14:
                break
            r += 1
            continue
        for c in range(1, 8):
            v = num_cells[c - 1]
            try:
                if v is not None and int(v) == day:
                    return r + 1, c, None   # celula de conteudo = linha de baixo
            except (ValueError, TypeError):
                continue
        r += 2
    return None, None, f"dia {day} nao achado no bloco de {names[0]}"

def resolve_merged(ws, row, col):
    """Se a celula esta num range merged, retorna a top-left (onde se escreve)."""
    coord = ws.cell(row, col).coordinate
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            return rng.min_row, rng.min_col
    return row, col

def write_item(ws, date_iso, titulo, apply=False):
    y, m, d = (int(x) for x in date_iso.split("-"))
    row, col, err = find_day_cell(ws, m, d)
    if err:
        print(f"  [SKIP] {date_iso} '{titulo[:40]}': {err}")
        return False
    row, col = resolve_merged(ws, row, col)
    cell = ws.cell(row, col)
    atual = str(cell.value or "").strip()
    marca = "Rax: " + titulo.strip()
    if marca in atual:
        print(f"  [DUP]  {date_iso} ja contem este item — pulando")
        return False
    novo = (atual + "\n" + marca) if atual else marca
    addr = f"{get_column_letter(col)}{row}"
    print(f"  [{'WRITE' if apply else 'DRY'}] {date_iso} -> celula {addr} | atual={atual[:30]!r} | +{marca[:50]!r}")
    if apply:
        cell.value = novo
    return True

def fetch_prod_raccoon():
    url = f"{PROD}/api/inbound/calendar?from=2026-01-01&to=2026-12-31"
    with urllib.request.urlopen(url, timeout=20) as r:
        data = json.loads(r.read().decode())
    return [p for p in data.get("posts", []) if p.get("fonte") == "raccoon"]

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date"); ap.add_argument("--titulo")
    ap.add_argument("--from-prod", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--path")
    a = ap.parse_args()

    src = resolve_path(a.path)
    if not src:
        sys.exit("xlsx da Duda nao encontrado no OneDrive.")
    print(f"[rax-xlsx] arquivo: {src}")
    print(f"[rax-xlsx] modo: {'APPLY (grava)' if a.apply else 'DRY-RUN (nao grava)'}")
    wb = load_workbook(src)
    ws = wb[wb.sheetnames[0]]

    itens = []
    if a.from_prod:
        itens = [(p["data"], p["titulo"]) for p in fetch_prod_raccoon()]
        print(f"[rax-xlsx] {len(itens)} itens fonte=raccoon do prod")
    elif a.date and a.titulo:
        itens = [(a.date, a.titulo)]
    else:
        sys.exit("informe --date e --titulo, ou --from-prod")

    n = sum(1 for dt, tt in itens if write_item(ws, dt, tt, apply=a.apply))
    if a.apply and n > 0:
        import shutil
        bak = src + ".bak-" + datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
        shutil.copy2(src, bak)
        print(f"[rax-xlsx] backup: {bak}")
        wb.save(src)
        print(f"[rax-xlsx] OK — {n} item(ns) gravado(s) no xlsx da Duda.")
    else:
        print(f"[rax-xlsx] dry-run: {n} item(ns) seriam gravados. Rode com --apply pra valer.")

if __name__ == "__main__":
    main()
