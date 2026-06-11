# -*- coding: utf-8 -*-
"""Espelha o Mural de Ideias -> planilha OneDrive (auto-sync SharePoint via OneDrive).
Lê /api/ideias (prod por padrão) e ANEXA no xlsx só as ideias novas (dedup por título).
Preserva edições manuais e formatação (openpyxl append). RODA LOCAL (tem acesso ao OneDrive).

Uso: python scripts/sync/sync_ideias_to_xlsx.py [local|railway]
"""
import sys, requests
from openpyxl import load_workbook

XLSX = r'C:\Users\Ruds\OneDrive - EPI USE BRASIL SERVIÇOS EM SISTEMAS LTDA\MARKETING\Planejamento\Repositório de Ideias de Marketing.xlsx'
TARGETS = {'local':'http://localhost:3000', 'railway':'https://epiuse-voices-optimizer.up.railway.app'}
base = TARGETS.get(sys.argv[1] if len(sys.argv) > 1 else 'railway', TARGETS['railway'])

# 1 — pega ideias da API
try:
    ideias = requests.get(base + '/api/ideias', timeout=20).json().get('ideias', [])
except Exception as e:
    print(f'ERRO ao ler {base}/api/ideias: {e}'); sys.exit(1)
print(f'{len(ideias)} ideias na API ({base})')

# 2 — abre xlsx, acha aba + coluna Tema
wb = load_workbook(XLSX)
ws = wb['Plan1'] if 'Plan1' in wb.sheetnames else wb.active

# header na linha 1: Tema | Ideias | Mente brilhante (+ extras que adicionamos)
headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    if v: headers[str(v).strip().lower()] = c
col_tema = headers.get('tema', 1)
col_ideia = headers.get('ideias', 2)
col_autor = headers.get('mente brilhante', 3)
# colunas extras (adiciona no header se não existirem)
def ensure_col(name):
    key = name.lower()
    if key in headers: return headers[key]
    c = ws.max_column + 1
    ws.cell(row=1, column=c, value=name)
    headers[key] = c
    return c
col_cat = ensure_col('Categoria')
col_imp = ensure_col('Impacto')
col_votos = ensure_col('Votos')
col_data = ensure_col('Cadastrada em')

# 3 — títulos já existentes (dedup)
existentes = set()
for r in range(2, ws.max_row + 1):
    t = ws.cell(row=r, column=col_tema).value
    if t and str(t).strip(): existentes.add(str(t).strip())

# 4 — anexa novas
novas = 0
for it in ideias:
    titulo = (it.get('titulo') or '').strip()
    if not titulo or titulo in existentes: continue
    r = ws.max_row + 1
    ws.cell(row=r, column=col_tema, value=titulo)
    ws.cell(row=r, column=col_ideia, value=it.get('descricao') or '')
    ws.cell(row=r, column=col_autor, value=it.get('autor') or 'Anônimo')
    ws.cell(row=r, column=col_cat, value=it.get('categoria') or '')
    ws.cell(row=r, column=col_imp, value=it.get('impacto') or '')
    ws.cell(row=r, column=col_votos, value=it.get('votos') or 0)
    ws.cell(row=r, column=col_data, value=(it.get('created_at') or '')[:10])
    existentes.add(titulo)
    novas += 1

wb.save(XLSX)
print(f'{novas} ideias novas anexadas na planilha. OneDrive vai sincronizar pro SharePoint.')
